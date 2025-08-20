"""
Real-Time Analytics Engine - 40by6
Provides real-time analytics, insights, and executive reporting
"""

import asyncio
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple, Set
from dataclasses import dataclass, field
from collections import defaultdict, deque
import pandas as pd
import numpy as np
from scipy import stats
import json
import redis.asyncio as redis
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import os

logger = logging.getLogger(__name__)


@dataclass
class MetricSnapshot:
    """Point-in-time metric snapshot"""
    timestamp: datetime
    metric_name: str
    value: float
    dimensions: Dict[str, str] = field(default_factory=dict)
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'timestamp': self.timestamp.isoformat(),
            'metric_name': self.metric_name,
            'value': self.value,
            'dimensions': self.dimensions
        }


@dataclass
class AnalyticsResult:
    """Result of an analytics query"""
    query_id: str
    timestamp: datetime
    data: Any
    metadata: Dict[str, Any] = field(default_factory=dict)
    visualizations: List[Dict[str, Any]] = field(default_factory=list)


class RealTimeMetricsCollector:
    """Collects and processes real-time metrics"""
    
    def __init__(self, redis_client: redis.Redis, window_size: int = 3600):
        self.redis = redis_client
        self.window_size = window_size  # seconds
        self.metrics_buffer = defaultdict(lambda: deque(maxlen=1000))
        self.aggregated_metrics = {}
        
    async def record_metric(self, metric: MetricSnapshot):
        """Record a metric data point"""
        # Add to buffer
        key = f"{metric.metric_name}:{json.dumps(metric.dimensions, sort_keys=True)}"
        self.metrics_buffer[key].append(metric)
        
        # Push to Redis for persistence
        await self.redis.zadd(
            f"metrics:{metric.metric_name}",
            {json.dumps(metric.to_dict()): metric.timestamp.timestamp()}
        )
        
        # Trim old data
        cutoff = datetime.utcnow().timestamp() - self.window_size
        await self.redis.zremrangebyscore(f"metrics:{metric.metric_name}", 0, cutoff)
    
    async def get_metric_series(
        self,
        metric_name: str,
        start_time: datetime,
        end_time: datetime,
        dimensions: Optional[Dict[str, str]] = None
    ) -> pd.DataFrame:
        """Get time series data for a metric"""
        # Fetch from Redis
        data = await self.redis.zrangebyscore(
            f"metrics:{metric_name}",
            start_time.timestamp(),
            end_time.timestamp()
        )
        
        # Parse and filter
        records = []
        for item in data:
            record = json.loads(item)
            
            # Filter by dimensions if specified
            if dimensions:
                match = all(
                    record['dimensions'].get(k) == v
                    for k, v in dimensions.items()
                )
                if not match:
                    continue
            
            records.append(record)
        
        # Convert to DataFrame
        if records:
            df = pd.DataFrame(records)
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            df = df.set_index('timestamp').sort_index()
            return df
        else:
            return pd.DataFrame()
    
    def calculate_aggregates(self, series: pd.DataFrame) -> Dict[str, float]:
        """Calculate aggregate statistics for a metric series"""
        if series.empty:
            return {}
        
        values = series['value']
        
        return {
            'count': len(values),
            'mean': values.mean(),
            'median': values.median(),
            'std': values.std(),
            'min': values.min(),
            'max': values.max(),
            'p95': values.quantile(0.95),
            'p99': values.quantile(0.99),
            'trend': self._calculate_trend(values)
        }
    
    def _calculate_trend(self, values: pd.Series) -> str:
        """Calculate trend direction"""
        if len(values) < 2:
            return 'stable'
        
        # Simple linear regression
        x = np.arange(len(values))
        slope, _, _, p_value, _ = stats.linregress(x, values)
        
        if p_value > 0.05:  # Not statistically significant
            return 'stable'
        elif slope > 0:
            return 'increasing'
        else:
            return 'decreasing'


class RealTimeAnalyticsEngine:
    """Main analytics engine for real-time insights"""
    
    def __init__(self, database_url: str = None, redis_url: str = None):
        self.database_url = database_url or os.getenv('DATABASE_URL')
        self.redis_url = redis_url or os.getenv('REDIS_URL', 'redis://localhost:6379')
        self.engine = create_engine(self.database_url)
        self.Session = sessionmaker(bind=self.engine)
        self.redis_client = None
        self.metrics_collector = None
        self.alert_thresholds = self._load_alert_thresholds()
        
    async def initialize(self):
        """Initialize the analytics engine"""
        self.redis_client = await redis.from_url(self.redis_url)
        self.metrics_collector = RealTimeMetricsCollector(self.redis_client)
        logger.info("Real-time analytics engine initialized")
    
    def _load_alert_thresholds(self) -> Dict[str, Dict[str, float]]:
        """Load alert thresholds configuration"""
        return {
            'scraper_failure_rate': {'warning': 0.2, 'critical': 0.5},
            'response_time_avg': {'warning': 10.0, 'critical': 30.0},
            'data_quality_score': {'warning': 0.7, 'critical': 0.5},
            'queue_size': {'warning': 100, 'critical': 500},
            'error_rate': {'warning': 0.1, 'critical': 0.3}
        }
    
    async def get_executive_dashboard(self) -> AnalyticsResult:
        """Generate executive dashboard with key metrics"""
        logger.info("Generating executive dashboard")
        
        now = datetime.utcnow()
        last_24h = now - timedelta(hours=24)
        last_7d = now - timedelta(days=7)
        last_30d = now - timedelta(days=30)
        
        # Gather all metrics
        metrics = await asyncio.gather(
            self._get_system_health_metrics(),
            self._get_scraper_performance_metrics(last_24h, now),
            self._get_data_quality_metrics(last_7d, now),
            self._get_legislative_activity_metrics(last_30d, now),
            self._get_cost_efficiency_metrics(last_30d, now)
        )
        
        system_health, scraper_perf, data_quality, legislative, cost = metrics
        
        # Create visualizations
        visualizations = [
            self._create_health_gauge(system_health),
            self._create_performance_timeline(scraper_perf),
            self._create_quality_heatmap(data_quality),
            self._create_activity_sunburst(legislative),
            self._create_cost_breakdown(cost)
        ]
        
        # Generate insights
        insights = self._generate_executive_insights(metrics)
        
        return AnalyticsResult(
            query_id='executive_dashboard',
            timestamp=now,
            data={
                'system_health': system_health,
                'scraper_performance': scraper_perf,
                'data_quality': data_quality,
                'legislative_activity': legislative,
                'cost_efficiency': cost,
                'insights': insights
            },
            metadata={
                'period': '24h/7d/30d',
                'last_updated': now.isoformat()
            },
            visualizations=visualizations
        )
    
    async def _get_system_health_metrics(self) -> Dict[str, Any]:
        """Get overall system health metrics"""
        with self.Session() as session:
            # Get scraper statistics
            scraper_stats = session.execute(text("""
                SELECT 
                    COUNT(*) as total_scrapers,
                    SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) as active_scrapers,
                    SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed_scrapers,
                    AVG(CASE WHEN last_run > NOW() - INTERVAL '24 hours' THEN 1 ELSE 0 END) as recent_run_rate
                FROM scrapers
            """)).fetchone()
            
            # Get recent error rate
            error_stats = session.execute(text("""
                SELECT 
                    COUNT(*) as total_runs,
                    SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed_runs
                FROM scraper_runs
                WHERE start_time > NOW() - INTERVAL '1 hour'
            """)).fetchone()
            
            error_rate = error_stats.failed_runs / error_stats.total_runs if error_stats.total_runs > 0 else 0
            
            # Calculate health score (0-100)
            health_score = (
                (scraper_stats.active_scrapers / scraper_stats.total_scrapers * 40) +
                ((1 - error_rate) * 30) +
                (scraper_stats.recent_run_rate * 30)
            )
            
            return {
                'health_score': health_score,
                'total_scrapers': scraper_stats.total_scrapers,
                'active_scrapers': scraper_stats.active_scrapers,
                'failed_scrapers': scraper_stats.failed_scrapers,
                'error_rate': error_rate,
                'status': self._get_health_status(health_score)
            }
    
    async def _get_scraper_performance_metrics(self, start_time: datetime, end_time: datetime) -> Dict[str, Any]:
        """Get scraper performance metrics"""
        with self.Session() as session:
            # Get performance data
            perf_data = session.execute(text("""
                SELECT 
                    DATE_TRUNC('hour', start_time) as hour,
                    AVG(EXTRACT(EPOCH FROM (end_time - start_time))) as avg_duration,
                    COUNT(*) as run_count,
                    SUM(records_scraped) as total_records,
                    AVG(data_quality_score) as avg_quality
                FROM scraper_runs
                WHERE start_time BETWEEN :start_time AND :end_time
                    AND end_time IS NOT NULL
                GROUP BY DATE_TRUNC('hour', start_time)
                ORDER BY hour
            """), {'start_time': start_time, 'end_time': end_time}).fetchall()
            
            # Get top performers
            top_performers = session.execute(text("""
                SELECT 
                    s.name,
                    COUNT(sr.id) as runs,
                    AVG(EXTRACT(EPOCH FROM (sr.end_time - sr.start_time))) as avg_duration,
                    SUM(sr.records_scraped) as total_records,
                    AVG(sr.data_quality_score) as quality_score
                FROM scrapers s
                JOIN scraper_runs sr ON s.id = sr.scraper_id
                WHERE sr.start_time BETWEEN :start_time AND :end_time
                    AND sr.status = 'completed'
                GROUP BY s.id, s.name
                ORDER BY total_records DESC
                LIMIT 10
            """), {'start_time': start_time, 'end_time': end_time}).fetchall()
            
            return {
                'timeline': [
                    {
                        'hour': row.hour.isoformat(),
                        'avg_duration': row.avg_duration,
                        'run_count': row.run_count,
                        'total_records': row.total_records,
                        'avg_quality': row.avg_quality
                    }
                    for row in perf_data
                ],
                'top_performers': [
                    {
                        'name': row.name,
                        'runs': row.runs,
                        'avg_duration': row.avg_duration,
                        'total_records': row.total_records,
                        'quality_score': row.quality_score
                    }
                    for row in top_performers
                ],
                'summary': {
                    'total_runs': sum(row.run_count for row in perf_data),
                    'total_records': sum(row.total_records for row in perf_data),
                    'avg_duration': np.mean([row.avg_duration for row in perf_data]) if perf_data else 0,
                    'avg_quality': np.mean([row.avg_quality for row in perf_data]) if perf_data else 0
                }
            }
    
    async def _get_data_quality_metrics(self, start_time: datetime, end_time: datetime) -> Dict[str, Any]:
        """Get data quality metrics"""
        with self.Session() as session:
            # Get quality trends
            quality_trends = session.execute(text("""
                SELECT 
                    DATE_TRUNC('day', created_at) as day,
                    issue_type,
                    severity,
                    COUNT(*) as issue_count
                FROM data_quality_issues
                WHERE created_at BETWEEN :start_time AND :end_time
                GROUP BY DATE_TRUNC('day', created_at), issue_type, severity
                ORDER BY day, issue_type
            """), {'start_time': start_time, 'end_time': end_time}).fetchall()
            
            # Get quality by category
            quality_by_category = session.execute(text("""
                SELECT 
                    s.category,
                    AVG(sr.data_quality_score) as avg_quality,
                    COUNT(DISTINCT s.id) as scraper_count
                FROM scrapers s
                JOIN scraper_runs sr ON s.id = sr.scraper_id
                WHERE sr.start_time BETWEEN :start_time AND :end_time
                    AND sr.data_quality_score IS NOT NULL
                GROUP BY s.category
            """), {'start_time': start_time, 'end_time': end_time}).fetchall()
            
            return {
                'trends': self._group_quality_trends(quality_trends),
                'by_category': [
                    {
                        'category': row.category,
                        'avg_quality': row.avg_quality,
                        'scraper_count': row.scraper_count
                    }
                    for row in quality_by_category
                ],
                'summary': {
                    'total_issues': sum(row.issue_count for row in quality_trends),
                    'critical_issues': sum(row.issue_count for row in quality_trends if row.severity == 'critical'),
                    'avg_quality_score': np.mean([row.avg_quality for row in quality_by_category]) if quality_by_category else 0
                }
            }
    
    async def _get_legislative_activity_metrics(self, start_time: datetime, end_time: datetime) -> Dict[str, Any]:
        """Get legislative activity metrics"""
        with self.Session() as session:
            # Get activity by jurisdiction
            activity_data = session.execute(text("""
                SELECT 
                    jurisdiction_type,
                    jurisdiction_name,
                    COUNT(DISTINCT CASE WHEN entity_type = 'bill' THEN entity_id END) as bills,
                    COUNT(DISTINCT CASE WHEN entity_type = 'vote' THEN entity_id END) as votes,
                    COUNT(DISTINCT CASE WHEN entity_type = 'committee' THEN entity_id END) as committees,
                    COUNT(DISTINCT CASE WHEN entity_type = 'event' THEN entity_id END) as events
                FROM legislative_activity
                WHERE created_at BETWEEN :start_time AND :end_time
                GROUP BY jurisdiction_type, jurisdiction_name
            """), {'start_time': start_time, 'end_time': end_time}).fetchall()
            
            # Get trending topics
            trending_topics = session.execute(text("""
                SELECT 
                    keyword,
                    COUNT(*) as mention_count,
                    COUNT(DISTINCT entity_id) as entity_count
                FROM keyword_mentions
                WHERE created_at BETWEEN :start_time AND :end_time
                GROUP BY keyword
                ORDER BY mention_count DESC
                LIMIT 20
            """), {'start_time': start_time, 'end_time': end_time}).fetchall()
            
            return {
                'by_jurisdiction': [
                    {
                        'type': row.jurisdiction_type,
                        'name': row.jurisdiction_name,
                        'bills': row.bills,
                        'votes': row.votes,
                        'committees': row.committees,
                        'events': row.events,
                        'total': row.bills + row.votes + row.committees + row.events
                    }
                    for row in activity_data
                ],
                'trending_topics': [
                    {
                        'keyword': row.keyword,
                        'mentions': row.mention_count,
                        'entities': row.entity_count
                    }
                    for row in trending_topics
                ],
                'summary': {
                    'total_bills': sum(row.bills for row in activity_data),
                    'total_votes': sum(row.votes for row in activity_data),
                    'total_committees': sum(row.committees for row in activity_data),
                    'total_events': sum(row.events for row in activity_data)
                }
            }
    
    async def _get_cost_efficiency_metrics(self, start_time: datetime, end_time: datetime) -> Dict[str, Any]:
        """Calculate cost efficiency metrics"""
        with self.Session() as session:
            # Get resource utilization
            resource_data = session.execute(text("""
                SELECT 
                    DATE_TRUNC('day', timestamp) as day,
                    AVG(cpu_usage) as avg_cpu,
                    AVG(memory_usage) as avg_memory,
                    SUM(api_calls) as total_api_calls,
                    SUM(data_transferred_gb) as data_gb
                FROM resource_metrics
                WHERE timestamp BETWEEN :start_time AND :end_time
                GROUP BY DATE_TRUNC('day', timestamp)
                ORDER BY day
            """), {'start_time': start_time, 'end_time': end_time}).fetchall()
            
            # Calculate costs (simplified)
            cpu_cost_per_hour = 0.05
            memory_cost_per_gb_hour = 0.01
            api_cost_per_1000 = 0.50
            bandwidth_cost_per_gb = 0.10
            
            total_cost = 0
            cost_breakdown = {
                'compute': 0,
                'memory': 0,
                'api_calls': 0,
                'bandwidth': 0
            }
            
            for row in resource_data:
                hours = 24  # Daily data
                cost_breakdown['compute'] += row.avg_cpu * cpu_cost_per_hour * hours
                cost_breakdown['memory'] += row.avg_memory * memory_cost_per_gb_hour * hours
                cost_breakdown['api_calls'] += (row.total_api_calls / 1000) * api_cost_per_1000
                cost_breakdown['bandwidth'] += row.data_gb * bandwidth_cost_per_gb
            
            total_cost = sum(cost_breakdown.values())
            
            # Get value metrics
            value_data = session.execute(text("""
                SELECT 
                    COUNT(DISTINCT entity_id) as unique_entities,
                    SUM(records_scraped) as total_records
                FROM scraper_runs
                WHERE start_time BETWEEN :start_time AND :end_time
                    AND status = 'completed'
            """), {'start_time': start_time, 'end_time': end_time}).fetchone()
            
            cost_per_entity = total_cost / value_data.unique_entities if value_data.unique_entities > 0 else 0
            cost_per_record = total_cost / value_data.total_records if value_data.total_records > 0 else 0
            
            return {
                'total_cost': total_cost,
                'cost_breakdown': cost_breakdown,
                'cost_per_entity': cost_per_entity,
                'cost_per_record': cost_per_record,
                'entities_collected': value_data.unique_entities,
                'records_collected': value_data.total_records,
                'efficiency_score': self._calculate_efficiency_score(total_cost, value_data.total_records)
            }
    
    def _get_health_status(self, score: float) -> str:
        """Determine health status from score"""
        if score >= 90:
            return 'excellent'
        elif score >= 75:
            return 'good'
        elif score >= 60:
            return 'fair'
        elif score >= 40:
            return 'poor'
        else:
            return 'critical'
    
    def _group_quality_trends(self, quality_trends) -> Dict[str, List[Dict[str, Any]]]:
        """Group quality trends by day"""
        grouped = defaultdict(lambda: defaultdict(int))
        
        for row in quality_trends:
            day_key = row.day.isoformat()
            issue_key = f"{row.issue_type}_{row.severity}"
            grouped[day_key][issue_key] = row.issue_count
        
        return [
            {
                'date': day,
                'issues': dict(issues)
            }
            for day, issues in grouped.items()
        ]
    
    def _calculate_efficiency_score(self, cost: float, records: int) -> float:
        """Calculate efficiency score (0-100)"""
        if records == 0:
            return 0
        
        # Target: $0.001 per record
        target_cost_per_record = 0.001
        actual_cost_per_record = cost / records
        
        # Score based on how close to target
        ratio = target_cost_per_record / actual_cost_per_record
        score = min(100, ratio * 100)
        
        return score
    
    def _create_health_gauge(self, health_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create health gauge visualization"""
        fig = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=health_data['health_score'],
            domain={'x': [0, 1], 'y': [0, 1]},
            title={'text': "System Health Score"},
            delta={'reference': 80, 'increasing': {'color': "green"}},
            gauge={
                'axis': {'range': [None, 100]},
                'bar': {'color': "darkblue"},
                'steps': [
                    {'range': [0, 40], 'color': "lightgray"},
                    {'range': [40, 60], 'color': "gray"},
                    {'range': [60, 80], 'color': "lightgreen"},
                    {'range': [80, 100], 'color': "green"}
                ],
                'threshold': {
                    'line': {'color': "red", 'width': 4},
                    'thickness': 0.75,
                    'value': 90
                }
            }
        ))
        
        fig.update_layout(height=300)
        
        return {
            'type': 'gauge',
            'data': fig.to_dict(),
            'title': 'System Health'
        }
    
    def _create_performance_timeline(self, perf_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create performance timeline visualization"""
        timeline = perf_data['timeline']
        
        if not timeline:
            return {'type': 'empty', 'message': 'No data available'}
        
        df = pd.DataFrame(timeline)
        
        # Create subplots
        fig = make_subplots(
            rows=2, cols=2,
            subplot_titles=('Run Count', 'Average Duration', 'Records Scraped', 'Quality Score'),
            specs=[[{'secondary_y': False}, {'secondary_y': False}],
                   [{'secondary_y': False}, {'secondary_y': False}]]
        )
        
        # Run count
        fig.add_trace(
            go.Scatter(x=df['hour'], y=df['run_count'], mode='lines+markers', name='Runs'),
            row=1, col=1
        )
        
        # Average duration
        fig.add_trace(
            go.Scatter(x=df['hour'], y=df['avg_duration'], mode='lines+markers', name='Duration (s)'),
            row=1, col=2
        )
        
        # Records scraped
        fig.add_trace(
            go.Bar(x=df['hour'], y=df['total_records'], name='Records'),
            row=2, col=1
        )
        
        # Quality score
        fig.add_trace(
            go.Scatter(x=df['hour'], y=df['avg_quality'], mode='lines+markers', name='Quality'),
            row=2, col=2
        )
        
        fig.update_layout(height=600, showlegend=False, title_text="24-Hour Performance Metrics")
        
        return {
            'type': 'timeline',
            'data': fig.to_dict(),
            'title': 'Performance Timeline'
        }
    
    def _create_quality_heatmap(self, quality_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create quality heatmap visualization"""
        by_category = quality_data['by_category']
        
        if not by_category:
            return {'type': 'empty', 'message': 'No quality data available'}
        
        categories = [item['category'] for item in by_category]
        quality_scores = [item['avg_quality'] for item in by_category]
        
        # Create color scale
        colors = ['red' if score < 0.6 else 'yellow' if score < 0.8 else 'green' for score in quality_scores]
        
        fig = go.Figure(data=[
            go.Bar(
                x=categories,
                y=quality_scores,
                marker_color=colors,
                text=[f"{score:.2f}" for score in quality_scores],
                textposition='auto'
            )
        ])
        
        fig.update_layout(
            title="Data Quality by Category",
            xaxis_title="Category",
            yaxis_title="Average Quality Score",
            yaxis_range=[0, 1],
            height=400
        )
        
        return {
            'type': 'bar',
            'data': fig.to_dict(),
            'title': 'Quality by Category'
        }
    
    def _create_activity_sunburst(self, activity_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create legislative activity sunburst chart"""
        by_jurisdiction = activity_data['by_jurisdiction']
        
        if not by_jurisdiction:
            return {'type': 'empty', 'message': 'No activity data available'}
        
        # Prepare hierarchical data
        labels = []
        parents = []
        values = []
        
        # Add root
        labels.append("Total Activity")
        parents.append("")
        values.append(sum(item['total'] for item in by_jurisdiction))
        
        # Group by type
        for jur_type in ['federal', 'provincial', 'municipal']:
            type_items = [item for item in by_jurisdiction if item['type'] == jur_type]
            if type_items:
                labels.append(jur_type.title())
                parents.append("Total Activity")
                values.append(sum(item['total'] for item in type_items))
                
                # Add individual jurisdictions
                for item in type_items:
                    labels.append(item['name'])
                    parents.append(jur_type.title())
                    values.append(item['total'])
        
        fig = go.Figure(go.Sunburst(
            labels=labels,
            parents=parents,
            values=values,
            branchvalues="total"
        ))
        
        fig.update_layout(
            title="Legislative Activity by Jurisdiction",
            height=500
        )
        
        return {
            'type': 'sunburst',
            'data': fig.to_dict(),
            'title': 'Activity Distribution'
        }
    
    def _create_cost_breakdown(self, cost_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create cost breakdown visualization"""
        breakdown = cost_data['cost_breakdown']
        
        fig = go.Figure(data=[
            go.Pie(
                labels=list(breakdown.keys()),
                values=list(breakdown.values()),
                hole=0.3
            )
        ])
        
        fig.update_layout(
            title=f"Cost Breakdown - Total: ${cost_data['total_cost']:.2f}",
            annotations=[{
                'text': f"${cost_data['cost_per_record']:.4f}<br>per record",
                'x': 0.5, 'y': 0.5,
                'font_size': 20,
                'showarrow': False
            }],
            height=400
        )
        
        return {
            'type': 'pie',
            'data': fig.to_dict(),
            'title': 'Cost Efficiency'
        }
    
    def _generate_executive_insights(self, metrics: List[Dict[str, Any]]) -> List[Dict[str, str]]:
        """Generate executive-level insights from metrics"""
        insights = []
        
        system_health, scraper_perf, data_quality, legislative, cost = metrics
        
        # System health insights
        if system_health['health_score'] < 70:
            insights.append({
                'type': 'warning',
                'category': 'system_health',
                'message': f"System health score is {system_health['health_score']:.0f}, below target of 80",
                'recommendation': "Investigate failed scrapers and high error rates"
            })
        
        # Performance insights
        if scraper_perf['summary']['avg_duration'] > 15:
            insights.append({
                'type': 'warning',
                'category': 'performance',
                'message': f"Average scraper duration is {scraper_perf['summary']['avg_duration']:.1f}s",
                'recommendation': "Optimize slow-running scrapers or increase parallelization"
            })
        
        # Data quality insights
        if data_quality['summary']['avg_quality_score'] < 0.8:
            insights.append({
                'type': 'warning',
                'category': 'data_quality',
                'message': f"Data quality score averaging {data_quality['summary']['avg_quality_score']:.2f}",
                'recommendation': "Review data validation rules and scraper logic"
            })
        
        # Cost efficiency insights
        efficiency_score = cost['efficiency_score']
        if efficiency_score < 70:
            insights.append({
                'type': 'info',
                'category': 'cost',
                'message': f"Cost efficiency at {efficiency_score:.0f}%, cost per record ${cost['cost_per_record']:.4f}",
                'recommendation': "Consider optimizing resource allocation and API usage"
            })
        
        # Positive insights
        if scraper_perf['summary']['total_records'] > 100000:
            insights.append({
                'type': 'success',
                'category': 'volume',
                'message': f"Collected {scraper_perf['summary']['total_records']:,} records in 24 hours",
                'recommendation': "Maintain current collection rate"
            })
        
        return insights
    
    async def generate_custom_report(self, report_config: Dict[str, Any]) -> AnalyticsResult:
        """Generate custom analytics report based on configuration"""
        logger.info(f"Generating custom report: {report_config.get('name', 'Unnamed')}")
        
        # Extract parameters
        start_time = pd.to_datetime(report_config.get('start_time', datetime.utcnow() - timedelta(days=7)))
        end_time = pd.to_datetime(report_config.get('end_time', datetime.utcnow()))
        metrics = report_config.get('metrics', ['all'])
        grouping = report_config.get('grouping', 'day')
        filters = report_config.get('filters', {})
        
        # Build and execute query
        query_results = await self._execute_custom_query(
            start_time, end_time, metrics, grouping, filters
        )
        
        # Generate visualizations
        visualizations = []
        for metric in metrics:
            viz = await self._create_custom_visualization(metric, query_results)
            if viz:
                visualizations.append(viz)
        
        return AnalyticsResult(
            query_id=f"custom_{datetime.utcnow().timestamp()}",
            timestamp=datetime.utcnow(),
            data=query_results,
            metadata={
                'config': report_config,
                'period': f"{start_time} to {end_time}",
                'record_count': len(query_results.get('data', []))
            },
            visualizations=visualizations
        )
    
    async def _execute_custom_query(
        self,
        start_time: datetime,
        end_time: datetime,
        metrics: List[str],
        grouping: str,
        filters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Execute custom analytics query"""
        # This would build dynamic SQL based on parameters
        # For now, returning sample data
        return {
            'data': [],
            'summary': {},
            'metadata': {
                'query_time': datetime.utcnow().isoformat(),
                'filters_applied': filters
            }
        }
    
    async def _create_custom_visualization(self, metric: str, data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Create visualization for custom metric"""
        # This would create appropriate visualization based on metric type
        return None
    
    async def stream_real_time_metrics(self, callback):
        """Stream real-time metrics to a callback function"""
        logger.info("Starting real-time metrics stream")
        
        while True:
            try:
                # Get latest metrics
                metrics = await self._get_current_metrics()
                
                # Check for alerts
                alerts = self._check_alert_conditions(metrics)
                
                # Call callback with metrics and alerts
                await callback({
                    'timestamp': datetime.utcnow().isoformat(),
                    'metrics': metrics,
                    'alerts': alerts
                })
                
                # Wait before next update
                await asyncio.sleep(5)  # Update every 5 seconds
                
            except Exception as e:
                logger.error(f"Error in metrics stream: {e}")
                await asyncio.sleep(10)
    
    async def _get_current_metrics(self) -> Dict[str, float]:
        """Get current real-time metrics"""
        with self.Session() as session:
            # Get current metrics
            metrics = {}
            
            # Active scrapers
            result = session.execute(text("""
                SELECT COUNT(*) as count
                FROM scraper_runs
                WHERE start_time > NOW() - INTERVAL '5 minutes'
                    AND end_time IS NULL
            """)).fetchone()
            metrics['active_scrapers'] = result.count
            
            # Recent error rate
            result = session.execute(text("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed
                FROM scraper_runs
                WHERE start_time > NOW() - INTERVAL '5 minutes'
            """)).fetchone()
            metrics['error_rate'] = result.failed / result.total if result.total > 0 else 0
            
            # Queue size (from Redis)
            if self.redis_client:
                queue_size = await self.redis_client.llen('scraper:queue')
                metrics['queue_size'] = queue_size
            
            # Average response time
            result = session.execute(text("""
                SELECT AVG(EXTRACT(EPOCH FROM (end_time - start_time))) as avg_time
                FROM scraper_runs
                WHERE start_time > NOW() - INTERVAL '5 minutes'
                    AND end_time IS NOT NULL
            """)).fetchone()
            metrics['avg_response_time'] = result.avg_time or 0
            
            return metrics
    
    def _check_alert_conditions(self, metrics: Dict[str, float]) -> List[Dict[str, Any]]:
        """Check if any metrics exceed alert thresholds"""
        alerts = []
        
        for metric_name, value in metrics.items():
            if metric_name in self.alert_thresholds:
                thresholds = self.alert_thresholds[metric_name]
                
                if value >= thresholds.get('critical', float('inf')):
                    alerts.append({
                        'severity': 'critical',
                        'metric': metric_name,
                        'value': value,
                        'threshold': thresholds['critical'],
                        'message': f"{metric_name} is critically high: {value:.2f}"
                    })
                elif value >= thresholds.get('warning', float('inf')):
                    alerts.append({
                        'severity': 'warning',
                        'metric': metric_name,
                        'value': value,
                        'threshold': thresholds['warning'],
                        'message': f"{metric_name} warning: {value:.2f}"
                    })
        
        return alerts


# Export functionality
class AnalyticsExporter:
    """Export analytics results in various formats"""
    
    @staticmethod
    async def export_to_pdf(result: AnalyticsResult, filename: str):
        """Export analytics result to PDF"""
        # This would use a PDF library like reportlab
        # For now, just save as JSON
        await AnalyticsExporter.export_to_json(result, filename.replace('.pdf', '.json'))
    
    @staticmethod
    async def export_to_excel(result: AnalyticsResult, filename: str):
        """Export analytics result to Excel"""
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        
        # Add metadata
        ws['A1'] = "Report Generated"
        ws['B1'] = result.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        ws['A2'] = "Query ID"
        ws['B2'] = result.query_id
        
        # Add data sheets based on content
        if isinstance(result.data, dict):
            for key, value in result.data.items():
                if isinstance(value, list) and value:
                    # Create new sheet
                    ws = wb.create_sheet(title=key[:30])  # Excel sheet name limit
                    
                    # Add headers
                    if isinstance(value[0], dict):
                        headers = list(value[0].keys())
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=header)
                        
                        # Add data
                        for row_idx, row_data in enumerate(value, 2):
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=row_idx, column=col, value=row_data.get(header))
        
        wb.save(filename)
    
    @staticmethod
    async def export_to_json(result: AnalyticsResult, filename: str):
        """Export analytics result to JSON"""
        import json
        
        data = {
            'query_id': result.query_id,
            'timestamp': result.timestamp.isoformat(),
            'data': result.data,
            'metadata': result.metadata,
            'visualizations': [
                {
                    'type': viz.get('type'),
                    'title': viz.get('title'),
                    'config': viz.get('data', {}).get('layout', {})
                }
                for viz in result.visualizations
            ]
        }
        
        with open(filename, 'w') as f:
            json.dump(data, f, indent=2, default=str)


# Example usage
async def run_analytics_demo():
    """Demo the analytics engine"""
    engine = RealTimeAnalyticsEngine()
    await engine.initialize()
    
    # Generate executive dashboard
    dashboard = await engine.get_executive_dashboard()
    
    # Export results
    exporter = AnalyticsExporter()
    await exporter.export_to_json(dashboard, 'executive_dashboard.json')
    await exporter.export_to_excel(dashboard, 'executive_dashboard.xlsx')
    
    logger.info("Analytics demo completed")
    
    return dashboard


if __name__ == "__main__":
    asyncio.run(run_analytics_demo())